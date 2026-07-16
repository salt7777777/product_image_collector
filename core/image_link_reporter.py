from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.models import ProductData, ImageItem
from utils.url_utils import get_url_ext
from utils.text_utils import safe_sku_name


class ImageLinkReportExporter:
    """
    商品图片链接总表导出器。

    生成一个按“图片维度”展开的 Excel 文件。
    """

    TYPE_LABELS = {
        "main": "主图",
        "detail": "详情图",
        "sku": "SKU图",
    }

    TYPE_DIRS = {
        "main": "主图",
        "detail": "详情图",
        "sku": "SKU图",
    }

    @staticmethod
    def save_image_link_report(
        base_dir: str,
        batch_items: list[dict],
        selected_types: dict[str, bool],
    ) -> Path:
        """
        保存商品图片链接总表。

        参数：
            base_dir:
                用户选择的输出根目录。

            batch_items:
                BatchDownloadWorker 中记录的商品下载明细。

            selected_types:
                用户选择的下载类型。
        """

        base_path = Path(base_dir)
        report_dir = base_path / "下载报告"
        report_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = report_dir / f"商品图片链接总表_{timestamp}.xlsx"

        wb = Workbook()

        ws_detail = wb.active
        ws_detail.title = "图片明细"

        ws_summary = wb.create_sheet("商品汇总")

        ImageLinkReportExporter._write_detail_sheet(
            ws=ws_detail,
            batch_items=batch_items,
            selected_types=selected_types,
        )

        ImageLinkReportExporter._write_summary_sheet(
            ws=ws_summary,
            batch_items=batch_items,
        )

        ImageLinkReportExporter._style_workbook(wb)

        wb.save(report_path)

        return report_path

    @staticmethod
    def _write_detail_sheet(
        ws,
        batch_items: list[dict],
        selected_types: dict[str, bool],
    ) -> None:
        """
        写入图片明细 Sheet。
        """

        headers = [
            "序号",
            "平台",
            "商品ID",
            "商品标题",
            "商品链接",
            "图片类型",
            "图片序号",
            "SKU名称",
            "图片来源",
            "图片URL",
            "是否选择下载",
            "预计保存目录",
            "预计文件名",
            "预计完整路径",
            "商品保存目录",
        ]

        ws.append(headers)

        row_index = 1

        for batch_item in batch_items:
            product: ProductData = batch_item["product"]
            product_dir = Path(batch_item["product_dir"])

            image_groups = {
                "main": product.main_images,
                "detail": product.detail_images,
                "sku": product.sku_images,
            }

            for image_type, images in image_groups.items():
                selected = bool(selected_types.get(image_type))
                type_label = ImageLinkReportExporter.TYPE_LABELS.get(image_type, image_type)
                type_dir_name = ImageLinkReportExporter.TYPE_DIRS.get(image_type, image_type)
                expected_dir = product_dir / type_dir_name

                for image_index, image_item in enumerate(images, start=1):
                    filename = ImageLinkReportExporter._build_expected_filename(
                        image_type=image_type,
                        index=image_index,
                        item=image_item,
                    )

                    expected_path = expected_dir / filename

                    ws.append(
                        [
                            row_index,
                            product.platform,
                            product.product_id,
                            product.title,
                            product.url,
                            type_label,
                            image_index,
                            image_item.sku_name or "",
                            image_item.source or "",
                            image_item.url,
                            "是" if selected else "否",
                            str(expected_dir),
                            filename,
                            str(expected_path),
                            str(product_dir),
                        ]
                    )

                    row_index += 1

    @staticmethod
    def _write_summary_sheet(
        ws,
        batch_items: list[dict],
    ) -> None:
        """
        写入商品汇总 Sheet。
        """

        headers = [
            "序号",
            "平台",
            "商品ID",
            "商品标题",
            "商品链接",
            "主图数",
            "详情图数",
            "SKU图数",
            "图片总数",
            "保存目录",
        ]

        ws.append(headers)

        for index, batch_item in enumerate(batch_items, start=1):
            product: ProductData = batch_item["product"]
            product_dir = batch_item["product_dir"]

            ws.append(
                [
                    index,
                    product.platform,
                    product.product_id,
                    product.title,
                    product.url,
                    len(product.main_images),
                    len(product.detail_images),
                    len(product.sku_images),
                    product.total_count(),
                    str(product_dir),
                ]
            )

    @staticmethod
    def _build_expected_filename(
        image_type: str,
        index: int,
        item: ImageItem,
    ) -> str:
        """
        构建预计保存文件名。

        逻辑尽量与 ImageDownloader._build_filename 保持一致。
        """

        ext = item.ext or get_url_ext(item.url)

        if image_type == "sku" and item.sku_name:
            sku_name = safe_sku_name(item.sku_name)
            return f"{index:03d}_{sku_name}.{ext}"

        return f"{index:03d}.{ext}"

    @staticmethod
    def _style_workbook(wb: Workbook) -> None:
        """
        美化 Excel。
        """

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin_side = Side(style="thin", color="D9E2F3")
        border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        for ws in wb.worksheets:
            ImageLinkReportExporter._style_worksheet(
                ws=ws,
                header_fill=header_fill,
                header_font=header_font,
                border=border,
            )

    @staticmethod
    def _style_worksheet(
        ws,
        header_fill,
        header_font,
        border,
    ) -> None:
        """
        美化单个 Sheet。
        """

        # 表头
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # 全表
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border

        # 自动筛选
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"

        # 冻结首行
        ws.freeze_panes = "A2"

        # 自动列宽
        for column_cells in ws.columns:
            column_letter = column_cells[0].column_letter

            max_length = 0

            for cell in column_cells:
                value = cell.value

                if value is None:
                    continue

                value = str(value)

                length = 0
                for ch in value:
                    if "\u4e00" <= ch <= "\u9fff":
                        length += 2
                    else:
                        length += 1

                max_length = max(max_length, length)

            width = min(max(max_length + 2, 10), 80)
            ws.column_dimensions[column_letter].width = width

        # 行高
        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 22
