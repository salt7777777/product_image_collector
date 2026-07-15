import json
from pathlib import Path
from datetime import datetime

from core.models import ProductData, DownloadResult


class TaskLogger:
    """
    生成采集日志、商品数据 JSON、批量下载报告、失败清单、Excel 报告。
    """

    @staticmethod
    def save_log(
        product_dir: Path,
        product: ProductData,
        selected_types: dict[str, bool],
        download_result: DownloadResult,
    ):
        log_path = product_dir / "采集日志.txt"

        lines = []
        lines.append("采集日志")
        lines.append("=" * 40)
        lines.append(f"采集时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"平台：{product.platform}")
        lines.append(f"商品ID：{product.product_id}")
        lines.append(f"商品标题：{product.title}")
        lines.append(f"商品链接：{product.url}")
        lines.append(f"保存路径：{product_dir}")
        lines.append("")
        lines.append("识别结果：")
        lines.append(f"主图：{len(product.main_images)} 张")
        lines.append(f"详情图：{len(product.detail_images)} 张")
        lines.append(f"SKU图：{len(product.sku_images)} 张")
        lines.append(f"总计：{product.total_count()} 张")
        lines.append("")
        lines.append("下载选择：")
        lines.append(f"主图：{'是' if selected_types.get('main') else '否'}")
        lines.append(f"详情图：{'是' if selected_types.get('detail') else '否'}")
        lines.append(f"SKU图：{'是' if selected_types.get('sku') else '否'}")
        lines.append("")
        lines.append("下载结果：")
        lines.append(f"计划下载：{download_result.total} 张")
        lines.append(f"成功下载：{download_result.success} 张")
        lines.append(f"失败下载：{download_result.failed} 张")
        lines.append(f"成功率：{download_result.success_rate}%")
        lines.append("")

        lines.append("图片去重：")
        lines.append(f"处理重复图片：{download_result.duplicate_removed} 张")
        lines.append(f"释放空间：{TaskLogger._format_bytes(download_result.duplicate_removed_bytes)}")
        lines.append("")

        if download_result.duplicate_items:
            lines.append("重复图片明细：")
            for index, item in enumerate(download_result.duplicate_items, start=1):
                lines.append(f"[{index}]")
                lines.append(f"    原始文件：{item.original_path}")
                lines.append(f"    重复文件：{item.duplicate_path}")
                lines.append(f"    MD5：{item.md5}")
                lines.append(f"    大小：{TaskLogger._format_bytes(item.size)}")
            lines.append("")
        else:
            lines.append("重复图片明细：无")
            lines.append("")

        lines.append("图片格式转换：")
        lines.append(f"转换成功：{download_result.converted_count} 张")
        lines.append(f"转换失败：{download_result.convert_failed} 张")
        lines.append("")

        if download_result.converted_items:
            lines.append("格式转换明细：")
            for index, item in enumerate(download_result.converted_items, start=1):
                lines.append(f"[{index}]")
                lines.append(f"    原始文件：{item.original_path}")
                lines.append(f"    输出文件：{item.output_path}")
                lines.append(f"    备份文件：{item.backup_path}")
                lines.append(f"    源格式：{item.source_format}")
                lines.append(f"    目标格式：{item.target_format}")
                lines.append(f"    状态：{'成功' if item.success else '失败'}")
                if item.reason:
                    lines.append(f"    原因：{item.reason}")
            lines.append("")
        else:
            lines.append("格式转换明细：无")
            lines.append("")

        if download_result.failed_items:
            lines.append("失败明细：")
            for index, item in enumerate(download_result.failed_items, start=1):
                lines.append(f"[{index}] 类型：{item.image_type}")
                lines.append(f"    文件名：{item.filename}")
                lines.append(f"    URL：{item.url}")
                lines.append(f"    原因：{item.reason}")
        else:
            lines.append("失败明细：无")

        log_path.write_text("\n".join(lines), encoding="utf-8")

    @staticmethod
    def save_product_json(
        product_dir: Path,
        product: ProductData,
        download_result: DownloadResult,
    ):
        json_path = product_dir / "商品数据.json"

        data = {
            "platform": product.platform,
            "product_id": product.product_id,
            "title": product.title,
            "url": product.url,
            "images": {
                "main": [img.url for img in product.main_images],
                "detail": [img.url for img in product.detail_images],
                "sku": [
                    {
                        "url": img.url,
                        "sku_name": img.sku_name,
                    }
                    for img in product.sku_images
                ],
            },
            "download_result": {
                "total": download_result.total,
                "success": download_result.success,
                "failed": download_result.failed,
                "success_rate": download_result.success_rate,
                "duplicate_removed": download_result.duplicate_removed,
                "duplicate_removed_bytes": download_result.duplicate_removed_bytes,
                "duplicate_items": [
                    {
                        "original_path": item.original_path,
                        "duplicate_path": item.duplicate_path,
                        "md5": item.md5,
                        "size": item.size,
                    }
                    for item in download_result.duplicate_items
                ],
                "converted_count": download_result.converted_count,
                "convert_failed": download_result.convert_failed,
                "converted_items": [
                    {
                        "original_path": item.original_path,
                        "output_path": item.output_path,
                        "backup_path": item.backup_path,
                        "source_format": item.source_format,
                        "target_format": item.target_format,
                        "success": item.success,
                        "reason": item.reason,
                    }
                    for item in download_result.converted_items
                ],
            },
        }

        json_path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    @staticmethod
    def save_batch_report(
        base_dir: str,
        batch_items: list[dict],
        aggregate_result: DownloadResult,
        selected_types: dict[str, bool],
        failed_parse_items: list[dict] | None = None,
    ) -> dict:
        """
        保存批量下载报告、失败清单和 Excel 报告。
        """

        base_path = Path(base_dir)
        report_dir = base_path / "下载报告"
        failed_dir = base_path / "失败清单"

        report_dir.mkdir(parents=True, exist_ok=True)
        failed_dir.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        report_path = report_dir / f"批量下载报告_{timestamp}.txt"
        excel_path = report_dir / f"批量下载报告_{timestamp}.xlsx"
        failed_path = failed_dir / f"失败清单_{timestamp}.txt"

        failed_parse_items = failed_parse_items or []

        # ------------------------------------------------------------
        # TXT 批量下载报告
        # ------------------------------------------------------------

        lines = []
        lines.append("批量下载报告")
        lines.append("=" * 60)
        lines.append(f"任务时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        lines.append(f"商品总数：{len(batch_items)}")
        lines.append(f"解析失败链接：{len(failed_parse_items)} 个")
        lines.append("")
        lines.append("下载选择：")
        lines.append(f"主图：{'是' if selected_types.get('main') else '否'}")
        lines.append(f"详情图：{'是' if selected_types.get('detail') else '否'}")
        lines.append(f"SKU图：{'是' if selected_types.get('sku') else '否'}")
        lines.append("")
        lines.append("下载汇总：")
        lines.append(f"计划下载：{aggregate_result.total} 张")
        lines.append(f"成功下载：{aggregate_result.success} 张")
        lines.append(f"失败下载：{aggregate_result.failed} 张")
        lines.append(f"成功率：{aggregate_result.success_rate}%")
        lines.append(f"去重处理：{aggregate_result.duplicate_removed} 张")
        lines.append(f"释放空间：{TaskLogger._format_bytes(aggregate_result.duplicate_removed_bytes)}")
        lines.append(f"格式转换成功：{aggregate_result.converted_count} 张")
        lines.append(f"格式转换失败：{aggregate_result.convert_failed} 张")
        lines.append("")
        lines.append("=" * 60)
        lines.append("商品明细")
        lines.append("=" * 60)
        lines.append("")

        for index, item in enumerate(batch_items, start=1):
            product = item["product"]
            product_dir = item["product_dir"]
            result = item["download_result"]

            lines.append(f"[{index}] {product.title}")
            lines.append(f"    平台：{product.platform}")
            lines.append(f"    商品ID：{product.product_id}")
            lines.append(f"    商品链接：{product.url}")
            lines.append(f"    保存路径：{product_dir}")
            lines.append("    识别结果：")
            lines.append(f"        主图：{len(product.main_images)} 张")
            lines.append(f"        详情图：{len(product.detail_images)} 张")
            lines.append(f"        SKU图：{len(product.sku_images)} 张")
            lines.append(f"        总计：{product.total_count()} 张")
            lines.append("    下载结果：")
            lines.append(f"        计划下载：{result.total} 张")
            lines.append(f"        成功下载：{result.success} 张")
            lines.append(f"        失败下载：{result.failed} 张")
            lines.append(f"        成功率：{result.success_rate}%")
            lines.append(f"        去重处理：{result.duplicate_removed} 张")
            lines.append(f"        释放空间：{TaskLogger._format_bytes(result.duplicate_removed_bytes)}")
            lines.append(f"        格式转换成功：{result.converted_count} 张")
            lines.append(f"        格式转换失败：{result.convert_failed} 张")
            lines.append("")

        report_path.write_text("\n".join(lines), encoding="utf-8")

        # ------------------------------------------------------------
        # 失败清单 TXT
        # ------------------------------------------------------------

        failed_lines = []
        failed_lines.append("失败清单")
        failed_lines.append("=" * 60)
        failed_lines.append(f"任务时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        failed_lines.append("")

        has_failed = False

        if failed_parse_items:
            has_failed = True
            failed_lines.append("一、解析失败链接")
            failed_lines.append("-" * 60)

            for index, item in enumerate(failed_parse_items, start=1):
                failed_lines.append(f"[{index}]")
                failed_lines.append(f"链接：{item.get('url', '')}")
                failed_lines.append(f"原因：{item.get('reason', '')}")
                failed_lines.append("")

        download_failed_count = 0

        for batch_item in batch_items:
            result = batch_item["download_result"]

            if result.failed_items:
                has_failed = True
                download_failed_count += len(result.failed_items)

        if download_failed_count:
            failed_lines.append("二、下载失败图片")
            failed_lines.append("-" * 60)

            item_index = 1

            for batch_item in batch_items:
                product = batch_item["product"]
                result = batch_item["download_result"]

                if not result.failed_items:
                    continue

                failed_lines.append(f"商品：{product.title}")
                failed_lines.append(f"平台：{product.platform}")
                failed_lines.append(f"商品ID：{product.product_id}")
                failed_lines.append(f"商品链接：{product.url}")

                for failed in result.failed_items:
                    failed_lines.append(f"    [{item_index}]")
                    failed_lines.append(f"    类型：{failed.image_type}")
                    failed_lines.append(f"    文件名：{failed.filename}")
                    failed_lines.append(f"    URL：{failed.url}")
                    failed_lines.append(f"    原因：{failed.reason}")
                    failed_lines.append("")
                    item_index += 1

        duplicate_count = 0

        for batch_item in batch_items:
            result = batch_item["download_result"]
            duplicate_count += len(result.duplicate_items)

        if duplicate_count:
            has_failed = True
            failed_lines.append("三、MD5重复图片处理记录")
            failed_lines.append("-" * 60)

            item_index = 1

            for batch_item in batch_items:
                product = batch_item["product"]
                result = batch_item["download_result"]

                if not result.duplicate_items:
                    continue

                failed_lines.append(f"商品：{product.title}")
                failed_lines.append(f"平台：{product.platform}")
                failed_lines.append(f"商品ID：{product.product_id}")
                failed_lines.append(f"商品链接：{product.url}")

                for duplicate in result.duplicate_items:
                    failed_lines.append(f"    [{item_index}]")
                    failed_lines.append(f"    原始文件：{duplicate.original_path}")
                    failed_lines.append(f"    重复文件：{duplicate.duplicate_path}")
                    failed_lines.append(f"    MD5：{duplicate.md5}")
                    failed_lines.append(f"    大小：{TaskLogger._format_bytes(duplicate.size)}")
                    failed_lines.append("")
                    item_index += 1

        convert_failed_count = 0

        for batch_item in batch_items:
            result = batch_item["download_result"]
            convert_failed_count += result.convert_failed

        if convert_failed_count:
            has_failed = True
            failed_lines.append("四、图片格式转换失败记录")
            failed_lines.append("-" * 60)

            item_index = 1

            for batch_item in batch_items:
                product = batch_item["product"]
                result = batch_item["download_result"]

                failed_convert_items = [
                    item for item in result.converted_items if not item.success
                ]

                if not failed_convert_items:
                    continue

                failed_lines.append(f"商品：{product.title}")
                failed_lines.append(f"平台：{product.platform}")
                failed_lines.append(f"商品ID：{product.product_id}")
                failed_lines.append(f"商品链接：{product.url}")

                for item in failed_convert_items:
                    failed_lines.append(f"    [{item_index}]")
                    failed_lines.append(f"    原始文件：{item.original_path}")
                    failed_lines.append(f"    源格式：{item.source_format}")
                    failed_lines.append(f"    目标格式：{item.target_format}")
                    failed_lines.append(f"    原因：{item.reason}")
                    failed_lines.append("")
                    item_index += 1

        if not has_failed:
            failed_lines.append("无失败链接、失败图片、重复图片或格式转换失败记录。")

        failed_path.write_text("\n".join(failed_lines), encoding="utf-8")

        # ------------------------------------------------------------
        # Excel 报告
        # ------------------------------------------------------------

        excel_created = False
        excel_error = ""

        try:
            TaskLogger._save_excel_report(
                excel_path=excel_path,
                batch_items=batch_items,
                aggregate_result=aggregate_result,
                selected_types=selected_types,
                failed_parse_items=failed_parse_items,
            )
            excel_created = True
        except Exception as e:
            excel_error = str(e)

            error_log_path = report_dir / f"Excel报告生成失败_{timestamp}.txt"
            error_log_path.write_text(
                f"Excel 报告生成失败：{excel_error}",
                encoding="utf-8",
            )

        return {
            "report_path": report_path,
            "failed_path": failed_path,
            "excel_path": excel_path if excel_created else None,
            "excel_error": excel_error,
        }

    @staticmethod
    def _save_excel_report(
        excel_path: Path,
        batch_items: list[dict],
        aggregate_result: DownloadResult,
        selected_types: dict[str, bool],
        failed_parse_items: list[dict] | None = None,
    ) -> None:
        """
        保存 Excel 批量报告。

        Sheet:
            1. 商品汇总
            2. 失败明细
            3. MD5去重记录
            4. 格式转换记录
        """

        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        failed_parse_items = failed_parse_items or []

        wb = Workbook()

        ws_summary = wb.active
        ws_summary.title = "商品汇总"

        ws_failed = wb.create_sheet("失败明细")
        ws_duplicate = wb.create_sheet("MD5去重记录")
        ws_convert = wb.create_sheet("格式转换记录")

        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(color="FFFFFF", bold=True)
        thin_side = Side(style="thin", color="D9E2F3")
        border = Border(
            left=thin_side,
            right=thin_side,
            top=thin_side,
            bottom=thin_side,
        )

        # ------------------------------------------------------------
        # Sheet1 商品汇总
        # ------------------------------------------------------------

        summary_headers = [
            "序号",
            "平台",
            "商品ID",
            "商品标题",
            "商品链接",
            "主图数",
            "详情图数",
            "SKU图数",
            "图片总数",
            "计划下载",
            "成功下载",
            "失败下载",
            "成功率",
            "MD5去重处理",
            "释放空间",
            "格式转换成功",
            "格式转换失败",
            "保存路径",
        ]

        ws_summary.append(["批量下载报告"])
        ws_summary.append([f"生成时间：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
        ws_summary.append([f"商品数量：{len(batch_items)}"])
        ws_summary.append([f"解析失败链接：{len(failed_parse_items)}"])
        ws_summary.append(
            [
                f"下载选择：主图={'是' if selected_types.get('main') else '否'}，"
                f"详情图={'是' if selected_types.get('detail') else '否'}，"
                f"SKU图={'是' if selected_types.get('sku') else '否'}"
            ]
        )
        ws_summary.append(
            [
                f"汇总：计划 {aggregate_result.total} 张，"
                f"成功 {aggregate_result.success} 张，"
                f"失败 {aggregate_result.failed} 张，"
                f"成功率 {aggregate_result.success_rate}%，"
                f"MD5去重处理 {aggregate_result.duplicate_removed} 张，"
                f"格式转换成功 {aggregate_result.converted_count} 张，"
                f"格式转换失败 {aggregate_result.convert_failed} 张"
            ]
        )
        ws_summary.append([])
        ws_summary.append(summary_headers)

        for index, item in enumerate(batch_items, start=1):
            product = item["product"]
            product_dir = item["product_dir"]
            result = item["download_result"]

            ws_summary.append(
                [
                    index,
                    product.platform,
                    product.product_id,
                    product.title,
                    product.url,
                    len(product.main_images),
                    len(product.detail_images),
                    len(product.sku_images),
                    product.total_count(),
                    result.total,
                    result.success,
                    result.failed,
                    f"{result.success_rate}%",
                    result.duplicate_removed,
                    TaskLogger._format_bytes(result.duplicate_removed_bytes),
                    result.converted_count,
                    result.convert_failed,
                    str(product_dir),
                ]
            )

        TaskLogger._style_worksheet(
            ws_summary,
            header_row=8,
            header_fill=header_fill,
            header_font=header_font,
            border=border,
        )

        # ------------------------------------------------------------
        # Sheet2 失败明细
        # ------------------------------------------------------------

        failed_headers = [
            "序号",
            "失败类型",
            "商品标题",
            "平台",
            "商品ID",
            "图片类型",
            "文件名",
            "链接或图片URL",
            "失败原因",
        ]

        ws_failed.append(failed_headers)

        failed_index = 1

        for item in failed_parse_items:
            ws_failed.append(
                [
                    failed_index,
                    "解析失败",
                    "",
                    "",
                    "",
                    "",
                    "",
                    item.get("url", ""),
                    item.get("reason", ""),
                ]
            )
            failed_index += 1

        for batch_item in batch_items:
            product = batch_item["product"]
            result = batch_item["download_result"]

            for failed in result.failed_items:
                ws_failed.append(
                    [
                        failed_index,
                        "下载失败",
                        product.title,
                        product.platform,
                        product.product_id,
                        failed.image_type,
                        failed.filename,
                        failed.url,
                        failed.reason,
                    ]
                )
                failed_index += 1

        if failed_index == 1:
            ws_failed.append(
                [
                    1,
                    "无",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "无解析失败链接或下载失败图片",
                ]
            )

        TaskLogger._style_worksheet(
            ws_failed,
            header_row=1,
            header_fill=header_fill,
            header_font=header_font,
            border=border,
        )

        # ------------------------------------------------------------
        # Sheet3 MD5去重记录
        # ------------------------------------------------------------

        duplicate_headers = [
            "序号",
            "商品标题",
            "平台",
            "商品ID",
            "原始文件",
            "重复文件",
            "MD5",
            "文件大小",
        ]

        ws_duplicate.append(duplicate_headers)

        duplicate_index = 1

        for batch_item in batch_items:
            product = batch_item["product"]
            result = batch_item["download_result"]

            for duplicate in result.duplicate_items:
                ws_duplicate.append(
                    [
                        duplicate_index,
                        product.title,
                        product.platform,
                        product.product_id,
                        duplicate.original_path,
                        duplicate.duplicate_path,
                        duplicate.md5,
                        TaskLogger._format_bytes(duplicate.size),
                    ]
                )
                duplicate_index += 1

        if duplicate_index == 1:
            ws_duplicate.append(
                [
                    1,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "无 MD5 重复图片处理记录",
                ]
            )

        TaskLogger._style_worksheet(
            ws_duplicate,
            header_row=1,
            header_fill=header_fill,
            header_font=header_font,
            border=border,
        )

        # ------------------------------------------------------------
        # Sheet4 格式转换记录
        # ------------------------------------------------------------

        convert_headers = [
            "序号",
            "商品标题",
            "平台",
            "商品ID",
            "原始文件",
            "输出文件",
            "备份文件",
            "源格式",
            "目标格式",
            "状态",
            "原因",
        ]

        ws_convert.append(convert_headers)

        convert_index = 1

        for batch_item in batch_items:
            product = batch_item["product"]
            result = batch_item["download_result"]

            for item in result.converted_items:
                ws_convert.append(
                    [
                        convert_index,
                        product.title,
                        product.platform,
                        product.product_id,
                        item.original_path,
                        item.output_path,
                        item.backup_path,
                        item.source_format,
                        item.target_format,
                        "成功" if item.success else "失败",
                        item.reason,
                    ]
                )
                convert_index += 1

        if convert_index == 1:
            ws_convert.append(
                [
                    1,
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "",
                    "无",
                    "无图片格式转换记录",
                ]
            )

        TaskLogger._style_worksheet(
            ws_convert,
            header_row=1,
            header_fill=header_fill,
            header_font=header_font,
            border=border,
        )

        # 冻结窗格
        ws_summary.freeze_panes = "A9"
        ws_failed.freeze_panes = "A2"
        ws_duplicate.freeze_panes = "A2"
        ws_convert.freeze_panes = "A2"

        # 自动筛选
        if ws_summary.max_row >= 8:
            ws_summary.auto_filter.ref = (
                f"A8:{get_column_letter(ws_summary.max_column)}{ws_summary.max_row}"
            )

        if ws_failed.max_row >= 1:
            ws_failed.auto_filter.ref = (
                f"A1:{get_column_letter(ws_failed.max_column)}{ws_failed.max_row}"
            )

        if ws_duplicate.max_row >= 1:
            ws_duplicate.auto_filter.ref = (
                f"A1:{get_column_letter(ws_duplicate.max_column)}{ws_duplicate.max_row}"
            )

        if ws_convert.max_row >= 1:
            ws_convert.auto_filter.ref = (
                f"A1:{get_column_letter(ws_convert.max_column)}{ws_convert.max_row}"
            )

        wb.save(excel_path)

    @staticmethod
    def _style_worksheet(
        ws,
        header_row: int,
        header_fill,
        header_font,
        border,
    ) -> None:
        """
        美化 Excel 工作表。
        """

        from openpyxl.styles import Alignment

        # 表头样式
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # 全表样式
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="center", wrap_text=True)
                cell.border = border

        # 自动列宽
        for column_cells in ws.columns:
            column_letter = column_cells[0].column_letter

            max_length = 0

            for cell in column_cells:
                value = cell.value

                if value is None:
                    continue

                value = str(value)

                length = 0
                for ch in value:
                    if "\u4e00" <= ch <= "\u9fff":
                        length += 2
                    else:
                        length += 1

                max_length = max(max_length, length)

            width = min(max(max_length + 2, 10), 60)
            ws.column_dimensions[column_letter].width = width

        for row in ws.iter_rows():
            ws.row_dimensions[row[0].row].height = 22

    @staticmethod
    def _format_bytes(size: int) -> str:
        """
        格式化字节数。
        """
        try:
            size = int(size)
        except Exception:
            size = 0

        if size < 1024:
            return f"{size} B"

        if size < 1024 * 1024:
            return f"{round(size / 1024, 2)} KB"

        if size < 1024 * 1024 * 1024:
            return f"{round(size / 1024 / 1024, 2)} MB"

        return f"{round(size / 1024 / 1024 / 1024, 2)} GB"
