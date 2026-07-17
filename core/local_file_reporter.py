from __future__ import annotations

import hashlib
from pathlib import Path
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from PIL import Image

from core.models import ProductData
from core.file_manager import FileManager


class LocalFileReportExporter:
    """
    本地文件清单 Excel 导出器。

    作用：
        扫描下载完成后的商品目录，生成“最终真实存在于硬盘上的图片文件清单”。

    与“商品图片链接总表”的区别：
        商品图片链接总表 = 解析阶段预计下载哪些 URL
        本地文件清单 = 下载、去重、格式转换、小图过滤后，最终留下了哪些文件
    """

    IMAGE_SUFFIXES = {
        ".jpg",
        ".jpeg",
        ".png",
        ".webp",
        ".bmp",
        ".gif",
    }

    IMAGE_TYPE_DIRS = {
        "main": "主图",
        "detail": "详情图",
        "sku": "SKU图",
    }

    IMAGE_TYPE_NAMES = {
        "main": "主图",
        "detail": "详情图",
        "sku": "SKU图",
    }

    @classmethod
    def export(
        cls,
        records: list[dict[str, Any]],
        report_dir: str | Path,
    ) -> Path | None:
        """
        导出本地文件清单。

        参数：
            records:
                [
                    {
                        "product": ProductData,
                        "product_dir": Path 或 str
                    }
                ]

            report_dir:
                报告输出目录，一般是 output/下载报告

        返回：
            生成的 Excel 路径；如果没有任何文件，返回 None。
        """

        if not records:
            return None

        report_dir = Path(report_dir)
        report_dir.mkdir(parents=True, exist_ok=True)

        rows = []

        for record in records:
            product = record.get("product")
            product_dir = record.get("product_dir")

            if not product or not product_dir:
                continue

            product_dir = Path(product_dir)

            if not product_dir.exists():
                continue

            rows.extend(cls._scan_product_dir(product, product_dir))

        if not rows:
            return None

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = report_dir / f"本地文件清单_{timestamp}.xlsx"

        wb = Workbook()

        ws = wb.active
        ws.title = "本地文件清单"

        headers = [
            "序号",
            "平台",
            "平台名称",
            "商品ID",
            "商品标题",
            "商品链接",
            "图片类型",
            "图片类型目录",
            "文件名",
            "文件扩展名",
            "文件大小KB",
            "图片宽度",
            "图片高度",
            "MD5",
            "本地完整路径",
            "商品目录",
            "生成时间",
        ]

        ws.append(headers)

        for index, row in enumerate(rows, start=1):
            ws.append(
                [
                    index,
                    row.get("platform", ""),
                    row.get("platform_name", ""),
                    row.get("product_id", ""),
                    row.get("title", ""),
                    row.get("url", ""),
                    row.get("image_type_name", ""),
                    row.get("image_type_dir", ""),
                    row.get("file_name", ""),
                    row.get("file_ext", ""),
                    row.get("size_kb", ""),
                    row.get("width", ""),
                    row.get("height", ""),
                    row.get("md5", ""),
                    row.get("file_path", ""),
                    row.get("product_dir", ""),
                    row.get("generated_at", ""),
                ]
            )

        cls._format_sheet(ws)

        # 商品汇总 Sheet
        summary_ws = wb.create_sheet("商品汇总")
        cls._write_summary(summary_ws, records, rows)

        wb.save(output_path)

        return output_path

    @classmethod
    def _scan_product_dir(
        cls,
        product: ProductData,
        product_dir: Path,
    ) -> list[dict[str, Any]]:
        rows = []

        generated_at = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for image_type, dir_name in cls.IMAGE_TYPE_DIRS.items():
            type_dir = product_dir / dir_name

            if not type_dir.exists():
                continue

            files = cls._list_image_files(type_dir)

            for file_path in files:
                width, height = cls._get_image_size(file_path)
                size_kb = cls._get_file_size_kb(file_path)
                md5 = cls._get_file_md5(file_path)

                rows.append(
                    {
                        "platform": product.platform,
                        "platform_name": FileManager.get_platform_display_name(product.platform),
                        "product_id": product.product_id,
                        "title": product.title,
                        "url": product.url,
                        "image_type": image_type,
                        "image_type_name": cls.IMAGE_TYPE_NAMES.get(image_type, image_type),
                        "image_type_dir": dir_name,
                        "file_name": file_path.name,
                        "file_ext": file_path.suffix.lower(),
                        "size_kb": size_kb,
                        "width": width,
                        "height": height,
                        "md5": md5,
                        "file_path": str(file_path.resolve()),
                        "product_dir": str(product_dir.resolve()),
                        "generated_at": generated_at,
                    }
                )

        return rows

    @classmethod
    def _list_image_files(cls, folder: Path) -> list[Path]:
        files = []

        if not folder.exists():
            return files

        for path in folder.rglob("*"):
            if not path.is_file():
                continue

            if path.suffix.lower() not in cls.IMAGE_SUFFIXES:
                continue

            # 跳过备份/过滤目录里的文件
            parts = [p.lower() for p in path.parts]

            skip_keywords = [
                "_重复图片备份",
                "_格式转换备份",
                "_小图过滤",
            ]

            if any(keyword.lower() in parts for keyword in skip_keywords):
                continue

            files.append(path)

        return sorted(files, key=lambda p: str(p).lower())

    @staticmethod
    def _get_file_size_kb(path: Path) -> float:
        try:
            return round(path.stat().st_size / 1024, 2)
        except Exception:
            return 0.0

    @staticmethod
    def _get_image_size(path: Path) -> tuple[int, int]:
        try:
            with Image.open(path) as img:
                return img.width, img.height
        except Exception:
            return 0, 0

    @staticmethod
    def _get_file_md5(path: Path) -> str:
        try:
            md5 = hashlib.md5()

            with path.open("rb") as f:
                for chunk in iter(lambda: f.read(1024 * 1024), b""):
                    md5.update(chunk)

            return md5.hexdigest()
        except Exception:
            return ""

    @staticmethod
    def _format_sheet(ws):
        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )

        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        widths = {
            "A": 8,
            "B": 12,
            "C": 14,
            "D": 18,
            "E": 40,
            "F": 45,
            "G": 12,
            "H": 14,
            "I": 18,
            "J": 12,
            "K": 14,
            "L": 12,
            "M": 12,
            "N": 34,
            "O": 80,
            "P": 80,
            "Q": 22,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    @classmethod
    def _write_summary(
        cls,
        ws,
        records: list[dict[str, Any]],
        rows: list[dict[str, Any]],
    ):
        headers = [
            "序号",
            "平台",
            "平台名称",
            "商品ID",
            "商品标题",
            "商品链接",
            "商品目录",
            "主图文件数",
            "详情图文件数",
            "SKU图文件数",
            "本地文件总数",
        ]

        ws.append(headers)

        for index, record in enumerate(records, start=1):
            product = record.get("product")
            product_dir = record.get("product_dir")

            if not product:
                continue

            product_dir_str = str(Path(product_dir).resolve()) if product_dir else ""

            product_rows = [
                row for row in rows
                if row.get("product_id") == product.product_id
                and row.get("platform") == product.platform
                and row.get("product_dir") == product_dir_str
            ]

            main_count = sum(1 for row in product_rows if row.get("image_type") == "main")
            detail_count = sum(1 for row in product_rows if row.get("image_type") == "detail")
            sku_count = sum(1 for row in product_rows if row.get("image_type") == "sku")

            ws.append(
                [
                    index,
                    product.platform,
                    FileManager.get_platform_display_name(product.platform),
                    product.product_id,
                    product.title,
                    product.url,
                    product_dir_str,
                    main_count,
                    detail_count,
                    sku_count,
                    len(product_rows),
                ]
            )

        for cell in ws[1]:
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        column_widths = {
            "A": 8,
            "B": 12,
            "C": 14,
            "D": 18,
            "E": 40,
            "F": 45,
            "G": 80,
            "H": 14,
            "I": 14,
            "J": 14,
            "K": 14,
        }

        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
