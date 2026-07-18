import json
from dataclasses import asdict
from pathlib import Path
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from core.models import ProductData, ReviewItem


class ReviewReportExporter:
    """
    评价数据报告导出器。

    输出：
        商品目录/评价数据.json
        商品目录/评价数据.xlsx
    """

    @staticmethod
    def export(
        product: ProductData,
        product_dir: str | Path,
        reviews: list[ReviewItem],
    ) -> dict:
        product_dir = Path(product_dir)
        product_dir.mkdir(parents=True, exist_ok=True)

        json_path = product_dir / "评价数据.json"
        excel_path = product_dir / "评价数据.xlsx"

        ReviewReportExporter._save_json(
            path=json_path,
            product=product,
            reviews=reviews,
        )

        ReviewReportExporter._save_excel(
            path=excel_path,
            product=product,
            reviews=reviews,
        )

        return {
            "json_path": json_path,
            "excel_path": excel_path,
        }

    @staticmethod
    def _save_json(
        path: Path,
        product: ProductData,
        reviews: list[ReviewItem],
    ):
        data = {
            "generated_at": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "platform": product.platform,
            "product_id": product.product_id,
            "title": product.title,
            "url": product.url,
            "review_count": len(reviews),
            "reviews": [asdict(item) for item in reviews],
        }

        path.write_text(
            json.dumps(data, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    @staticmethod
    def _save_excel(
        path: Path,
        product: ProductData,
        reviews: list[ReviewItem],
    ):
        wb = Workbook()

        ws = wb.active
        ws.title = "评价明细"

        headers = [
            "序号",
            "平台",
            "商品ID",
            "商品标题",
            "商品链接",
            "用户昵称",
            "评价时间",
            "购买规格",
            "点赞数",
            "评价内容",
            "图片数量",
            "视频数量",
            "图片URL",
            "视频URL",
            "本地图片路径",
            "本地视频路径",
            "来源",
        ]

        ws.append(headers)

        for review in reviews:
            image_urls = "\n".join([m.url for m in review.images])
            video_urls = "\n".join([m.url for m in review.videos])

            local_image_paths = "\n".join(
                [m.local_path for m in review.images if m.local_path]
            )
            local_video_paths = "\n".join(
                [m.local_path for m in review.videos if m.local_path]
            )

            ws.append(
                [
                    review.index,
                    product.platform,
                    product.product_id,
                    product.title,
                    product.url,
                    review.user_name,
                    review.date,
                    review.sku_info,
                    review.like_count,
                    review.content,
                    len(review.images),
                    len(review.videos),
                    image_urls,
                    video_urls,
                    local_image_paths,
                    local_video_paths,
                    review.source,
                ]
            )

        ReviewReportExporter._style_sheet(ws)

        summary_ws = wb.create_sheet("商品汇总")
        ReviewReportExporter._write_summary(summary_ws, product, reviews)

        wb.save(path)

    @staticmethod
    def _style_sheet(ws):
        header_fill = PatternFill(
            fill_type="solid",
            fgColor="1F4E78",
        )
        header_font = Font(
            color="FFFFFF",
            bold=True,
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.freeze_panes = "A2"

        widths = {
            "A": 8,
            "B": 12,
            "C": 18,
            "D": 40,
            "E": 45,
            "F": 18,
            "G": 18,
            "H": 30,
            "I": 10,
            "J": 60,
            "K": 10,
            "L": 10,
            "M": 80,
            "N": 80,
            "O": 80,
            "P": 80,
            "Q": 20,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    @staticmethod
    def _write_summary(ws, product: ProductData, reviews: list[ReviewItem]):
        total_images = sum(len(r.images) for r in reviews)
        total_videos = sum(len(r.videos) for r in reviews)
        image_success = sum(
            1 for r in reviews for m in r.images if m.download_success
        )
        video_success = sum(
            1 for r in reviews for m in r.videos if m.download_success
        )

        rows = [
            ["生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["平台", product.platform],
            ["商品ID", product.product_id],
            ["商品标题", product.title],
            ["商品链接", product.url],
            ["评价条数", len(reviews)],
            ["评价图片总数", total_images],
            ["评价图片下载成功", image_success],
            ["评价视频总数", total_videos],
            ["评价视频下载成功", video_success],
        ]

        for row in rows:
            ws.append(row)

        for cell in ws[1]:
            cell.fill = PatternFill(fill_type="solid", fgColor="1F4E78")
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        ws.column_dimensions["A"].width = 24
        ws.column_dimensions["B"].width = 80

        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
