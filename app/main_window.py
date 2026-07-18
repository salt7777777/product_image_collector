from datetime import datetime
from pathlib import Path
from copy import deepcopy
import json
import os
import re
import shutil
from urllib.parse import urlparse, parse_qs

from PySide6.QtWidgets import (
    QMainWindow,
    QWidget,
    QLabel,
    QTextEdit,
    QLineEdit,
    QPushButton,
    QVBoxLayout,
    QHBoxLayout,
    QGroupBox,
    QFileDialog,
    QCheckBox,
    QProgressBar,
    QMessageBox,
    QSpinBox,
    QComboBox,
    QInputDialog,
)

from app.workers import (
    BatchParseWorker,
    BatchDownloadWorker,
    RetryFailedDownloadWorker,
)
from app.browser_session_worker import LoginBrowserWorker
from core.preview_generator import PreviewGenerator
from core.app_config import AppConfig
from core.task_state import TaskStateManager


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()

        self.config = AppConfig.load()

        self.setWindowTitle("商品图片采集工具")
        self.resize(980, 850)

        self.product = None
        self.products = []
        self.original_products = []
        self.failed_parse_items = []
        self.selection_file_path = None

        self.last_product_dir = None
        self.last_base_dir = None
        self.last_preview_path = None

        self.parse_worker = None
        self.download_worker = None
        self.retry_worker = None
        self.login_browser_worker = None

        self.last_retry_items = []

        self._init_ui()
        self._bind_events()
        self.refresh_resume_task_button()

    # ------------------------------------------------------------------
    # UI
    # ------------------------------------------------------------------

    def _init_ui(self):
        root = QWidget()
        self.setCentralWidget(root)

        main_layout = QVBoxLayout(root)

        input_group = QGroupBox("商品链接")
        input_layout = QVBoxLayout(input_group)

        self.url_input = QTextEdit()
        self.url_input.setPlaceholderText(
            "请粘贴商品链接，支持单个或批量链接。\n"
            "批量链接建议每行一个。\n"
            "支持导入 TXT / CSV / Excel 文件。\n"
            "长链接会自动简化。"
        )
        self.url_input.setFixedHeight(110)

        input_layout.addWidget(self.url_input)
        main_layout.addWidget(input_group)

        info_group = QGroupBox("商品信息")
        info_layout = QVBoxLayout(info_group)

        self.platform_label = QLabel("平台：-")
        self.product_id_label = QLabel("商品ID：-")
        self.title_label = QLabel("商品标题：-")

        info_layout.addWidget(self.platform_label)
        info_layout.addWidget(self.product_id_label)
        info_layout.addWidget(self.title_label)

        main_layout.addWidget(info_group)

        result_group = QGroupBox("识别结果与下载类型")
        result_layout = QVBoxLayout(result_group)

        row1 = QHBoxLayout()

        self.main_count_label = QLabel("主图：- 张")
        self.detail_count_label = QLabel("详情图：- 张")
        self.sku_count_label = QLabel("SKU图：- 张")

        row1.addWidget(self.main_count_label)
        row1.addWidget(self.detail_count_label)
        row1.addWidget(self.sku_count_label)
        row1.addStretch()

        row2 = QHBoxLayout()

        self.main_checkbox = QCheckBox("下载主图")
        self.detail_checkbox = QCheckBox("下载详情图")
        self.sku_checkbox = QCheckBox("下载SKU图")
        self.high_quality_checkbox = QCheckBox("尽量下载高清图")

        self.main_checkbox.setChecked(self.config.download_main)
        self.detail_checkbox.setChecked(self.config.download_detail)
        self.sku_checkbox.setChecked(self.config.download_sku)
        self.high_quality_checkbox.setChecked(self.config.high_quality)

        row2.addWidget(self.main_checkbox)
        row2.addWidget(self.detail_checkbox)
        row2.addWidget(self.sku_checkbox)
        row2.addWidget(self.high_quality_checkbox)
        row2.addStretch()

        result_layout.addLayout(row1)
        result_layout.addLayout(row2)

        main_layout.addWidget(result_group)

        settings_group = QGroupBox("高级设置")
        settings_layout = QVBoxLayout(settings_group)

        settings_row1 = QHBoxLayout()

        self.timeout_label = QLabel("下载超时(秒)：")
        self.timeout_spinbox = QSpinBox()
        self.timeout_spinbox.setRange(5, 120)
        self.timeout_spinbox.setValue(self.config.download_timeout)

        self.retries_label = QLabel("重试次数：")
        self.retries_spinbox = QSpinBox()
        self.retries_spinbox.setRange(0, 10)
        self.retries_spinbox.setValue(self.config.download_retries)

        self.login_wait_label = QLabel("登录等待(秒)：")
        self.login_wait_spinbox = QSpinBox()
        self.login_wait_spinbox.setRange(30, 600)
        self.login_wait_spinbox.setValue(self.config.login_wait_seconds)

        settings_row1.addWidget(self.timeout_label)
        settings_row1.addWidget(self.timeout_spinbox)
        settings_row1.addWidget(self.retries_label)
        settings_row1.addWidget(self.retries_spinbox)
        settings_row1.addWidget(self.login_wait_label)
        settings_row1.addWidget(self.login_wait_spinbox)
        settings_row1.addStretch()

        settings_row2 = QHBoxLayout()

        self.headless_checkbox = QCheckBox("隐藏浏览器窗口")
        self.headless_checkbox.setChecked(self.config.headless)

        self.organize_by_date_checkbox = QCheckBox("按日期分类保存")
        self.organize_by_date_checkbox.setChecked(self.config.organize_by_date)

        self.organize_by_platform_checkbox = QCheckBox("按平台分类保存")
        self.organize_by_platform_checkbox.setChecked(self.config.organize_by_platform)

        self.dedupe_images_checkbox = QCheckBox("下载后MD5去重")
        self.dedupe_images_checkbox.setChecked(self.config.dedupe_images)

        self.image_format_label = QLabel("输出格式：")
        self.image_format_combo = QComboBox()
        self.image_format_combo.addItem("保持原格式", "original")
        self.image_format_combo.addItem("全部转 JPG", "jpg")
        self.image_format_combo.addItem("全部转 PNG", "png")
        self.image_format_combo.addItem("全部转 WEBP", "webp")
        self._set_image_format_combo_value(self.config.image_output_format)

        settings_row2.addWidget(self.headless_checkbox)
        settings_row2.addWidget(self.organize_by_date_checkbox)
        settings_row2.addWidget(self.organize_by_platform_checkbox)
        settings_row2.addWidget(self.dedupe_images_checkbox)
        settings_row2.addWidget(self.image_format_label)
        settings_row2.addWidget(self.image_format_combo)
        settings_row2.addStretch()

        settings_row3 = QHBoxLayout()

        self.filter_small_images_checkbox = QCheckBox("过滤小图")
        self.filter_small_images_checkbox.setChecked(self.config.filter_small_images)

        self.min_width_label = QLabel("最小宽：")
        self.min_width_spinbox = QSpinBox()
        self.min_width_spinbox.setRange(1, 5000)
        self.min_width_spinbox.setValue(self.config.min_image_width)
        self.min_width_spinbox.setSuffix(" px")

        self.min_height_label = QLabel("最小高：")
        self.min_height_spinbox = QSpinBox()
        self.min_height_spinbox.setRange(1, 5000)
        self.min_height_spinbox.setValue(self.config.min_image_height)
        self.min_height_spinbox.setSuffix(" px")

        settings_row3.addWidget(self.filter_small_images_checkbox)
        settings_row3.addWidget(self.min_width_label)
        settings_row3.addWidget(self.min_width_spinbox)
        settings_row3.addWidget(self.min_height_label)
        settings_row3.addWidget(self.min_height_spinbox)
        settings_row3.addStretch()

        settings_row4 = QHBoxLayout()

        self.open_login_browser_button = QPushButton("打开登录浏览器")
        self.clear_browser_data_button = QPushButton("清除登录状态")

        settings_row4.addWidget(self.open_login_browser_button)
        settings_row4.addWidget(self.clear_browser_data_button)
        settings_row4.addStretch()

        settings_layout.addLayout(settings_row1)
        settings_layout.addLayout(settings_row2)
        settings_layout.addLayout(settings_row3)
        settings_layout.addLayout(settings_row4)

        main_layout.addWidget(settings_group)

        path_group = QGroupBox("保存路径")
        path_layout = QHBoxLayout(path_group)

        self.path_input = QLineEdit()
        self.path_input.setPlaceholderText("请选择图片保存路径")
        self.path_input.setText(self.config.save_dir)

        self.choose_path_button = QPushButton("选择目录")

        path_layout.addWidget(self.path_input)
        path_layout.addWidget(self.choose_path_button)

        main_layout.addWidget(path_group)

        button_layout = QHBoxLayout()

        self.parse_button = QPushButton("解析商品")
        self.download_button = QPushButton("开始下载")
        self.retry_failed_button = QPushButton("重试失败")
        self.stop_button = QPushButton("停止任务")
        self.preview_button = QPushButton("预览图片")
        self.import_selection_button = QPushButton("导入选择结果")
        self.open_dir_button = QPushButton("打开保存目录")
        self.import_links_button = QPushButton("导入链接文件")

        self.resume_task_button = QPushButton("继续上次任务")
        self.clear_log_button = QPushButton("清空日志")

        self.download_button.setEnabled(False)
        self.retry_failed_button.setEnabled(False)
        self.stop_button.setEnabled(False)
        self.preview_button.setEnabled(False)
        self.import_selection_button.setEnabled(False)
        self.open_dir_button.setEnabled(False)

        button_layout.addWidget(self.parse_button)
        button_layout.addWidget(self.download_button)
        button_layout.addWidget(self.retry_failed_button)
        button_layout.addWidget(self.stop_button)
        button_layout.addWidget(self.preview_button)
        button_layout.addWidget(self.import_selection_button)
        button_layout.addWidget(self.open_dir_button)
        button_layout.addWidget(self.import_links_button)

        button_layout.addWidget(self.resume_task_button)
        button_layout.addWidget(self.clear_log_button)
        button_layout.addStretch()

        main_layout.addLayout(button_layout)

        self.progress_bar = QProgressBar()
        self.progress_bar.setValue(0)
        main_layout.addWidget(self.progress_bar)

        log_group = QGroupBox("实时日志")
        log_layout = QVBoxLayout(log_group)

        self.log_text = QTextEdit()
        self.log_text.setReadOnly(True)

        log_layout.addWidget(self.log_text)
        main_layout.addWidget(log_group)

    def _bind_events(self):
        self.choose_path_button.clicked.connect(self.choose_path)
        self.parse_button.clicked.connect(self.parse_product)
        self.download_button.clicked.connect(self.download_images)
        self.retry_failed_button.clicked.connect(self.retry_failed_images)
        self.stop_button.clicked.connect(self.stop_current_task)
        self.preview_button.clicked.connect(self.preview_images)
        self.import_selection_button.clicked.connect(self.import_selection_file)
        
        self.resume_task_button.clicked.connect(self.resume_last_task)
        self.open_dir_button.clicked.connect(self.open_last_dir)
        self.import_links_button.clicked.connect(self.import_links_file)
        self.clear_log_button.clicked.connect(self.clear_log)
        self.open_login_browser_button.clicked.connect(self.open_login_browser)
        self.clear_browser_data_button.clicked.connect(self.clear_browser_data)

        self.path_input.textChanged.connect(self.save_current_config)

        self.main_checkbox.stateChanged.connect(self.save_current_config)
        self.detail_checkbox.stateChanged.connect(self.save_current_config)
        self.sku_checkbox.stateChanged.connect(self.save_current_config)
        self.high_quality_checkbox.stateChanged.connect(self.save_current_config)

        self.timeout_spinbox.valueChanged.connect(self.save_current_config)
        self.retries_spinbox.valueChanged.connect(self.save_current_config)
        self.login_wait_spinbox.valueChanged.connect(self.save_current_config)

        self.headless_checkbox.stateChanged.connect(self.save_current_config)
        self.organize_by_date_checkbox.stateChanged.connect(self.save_current_config)
        self.organize_by_platform_checkbox.stateChanged.connect(self.save_current_config)
        self.dedupe_images_checkbox.stateChanged.connect(self.save_current_config)
        self.image_format_combo.currentIndexChanged.connect(self.save_current_config)

        self.filter_small_images_checkbox.stateChanged.connect(self.save_current_config)
        self.min_width_spinbox.valueChanged.connect(self.save_current_config)
        self.min_height_spinbox.valueChanged.connect(self.save_current_config)

    def choose_path(self):
        path = QFileDialog.getExistingDirectory(self, "选择保存目录")
        if path:
            self.path_input.setText(path)

    def save_current_config(self):
        try:
            self.config.save_dir = self.path_input.text().strip() or str(Path("output").absolute())

            self.config.download_main = self.main_checkbox.isChecked()
            self.config.download_detail = self.detail_checkbox.isChecked()
            self.config.download_sku = self.sku_checkbox.isChecked()
            self.config.high_quality = self.high_quality_checkbox.isChecked()

            self.config.download_timeout = self.timeout_spinbox.value()
            self.config.download_retries = self.retries_spinbox.value()

            self.config.headless = self.headless_checkbox.isChecked()
            self.config.login_wait_seconds = self.login_wait_spinbox.value()

            self.config.organize_by_date = self.organize_by_date_checkbox.isChecked()
            self.config.organize_by_platform = self.organize_by_platform_checkbox.isChecked()
            self.config.dedupe_images = self.dedupe_images_checkbox.isChecked()

            self.config.image_output_format = self._get_image_format_combo_value()

            self.config.filter_small_images = self.filter_small_images_checkbox.isChecked()
            self.config.min_image_width = self.min_width_spinbox.value()
            self.config.min_image_height = self.min_height_spinbox.value()

            self.config.save()

        except Exception as e:
            try:
                self.log(f"保存配置失败：{e}")
            except Exception:
                pass

    # ------------------------------------------------------------------
    # 浏览器登录状态
    # ------------------------------------------------------------------

    def open_login_browser(self):
        if self.login_browser_worker and self.login_browser_worker.isRunning():
            self.log("正在结束登录浏览器，请稍候...")
            self.open_login_browser_button.setEnabled(False)

            try:
                self.login_browser_worker.stop()
            except Exception:
                pass

            return

        if self._is_task_running():
            QMessageBox.warning(
                self,
                "提示",
                "当前有解析或下载任务正在运行，请等待任务结束后再打开登录浏览器。",
            )
            return

        platform_items = [
            "淘宝",
            "天猫",
            "京东",
            "拼多多",
            "1688",
        ]

        platform_text, ok = QInputDialog.getItem(
            self,
            "选择登录平台",
            "请选择要打开的登录平台：",
            platform_items,
            0,
            False,
        )

        if not ok or not platform_text:
            return

        platform_map = {
            "淘宝": "taobao",
            "天猫": "tmall",
            "京东": "jd",
            "拼多多": "pdd",
            "1688": "1688",
        }

        platform = platform_map.get(platform_text)

        if not platform:
            QMessageBox.warning(self, "提示", "未知平台。")
            return

        self.login_browser_worker = LoginBrowserWorker(platform=platform)
        self.login_browser_worker.log_signal.connect(self.log)
        self.login_browser_worker.error_signal.connect(self.on_login_browser_error)
        self.login_browser_worker.finished_signal.connect(self.on_login_browser_finished)

        self.open_login_browser_button.setText("结束登录浏览器")
        self.open_login_browser_button.setEnabled(True)
        self.clear_browser_data_button.setEnabled(False)

        self.login_browser_worker.start()

    def on_login_browser_error(self, message: str):
        self.log(message)
        QMessageBox.critical(self, "登录浏览器错误", message)

    def on_login_browser_finished(self):
        self.open_login_browser_button.setText("打开登录浏览器")
        self.open_login_browser_button.setEnabled(True)
        self.clear_browser_data_button.setEnabled(True)

        self.log("如已完成登录，后续解析商品时将复用当前登录状态。")

        try:
            self.login_browser_worker = None
        except Exception:
            pass

    def clear_browser_data(self):
        if self._is_task_running():
            QMessageBox.warning(
                self,
                "提示",
                "当前有解析或下载任务正在运行，请等待任务结束后再清除登录状态。",
            )
            return

        if self.login_browser_worker and self.login_browser_worker.isRunning():
            QMessageBox.warning(
                self,
                "提示",
                "登录浏览器正在运行，请先点击“结束登录浏览器”，等待按钮恢复后再清除登录状态。",
            )
            return

        browser_data_dir = Path("browser_data")

        if not browser_data_dir.exists():
            QMessageBox.information(self, "提示", "当前没有 browser_data 登录状态目录。")
            return

        reply = QMessageBox.question(
            self,
            "确认清除登录状态",
            "确定要清除浏览器登录状态吗？\n\n"
            "这会删除 browser_data 目录。\n"
            "清除后淘宝、天猫、京东、拼多多可能需要重新登录。",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.No,
        )

        if reply != QMessageBox.Yes:
            return

        try:
            shutil.rmtree(browser_data_dir)
            self.log("浏览器登录状态已清除：browser_data")
            QMessageBox.information(self, "完成", "浏览器登录状态已清除。")

        except Exception as e:
            QMessageBox.critical(
                self,
                "清除失败",
                f"清除 browser_data 失败：{e}\n\n"
                "请确认没有浏览器窗口正在使用该目录。",
            )

    def _is_task_running(self) -> bool:
        if self.parse_worker and self.parse_worker.isRunning():
            return True

        if self.download_worker and self.download_worker.isRunning():
            return True

        if self.retry_worker and self.retry_worker.isRunning():
            return True

        return False

    # ------------------------------------------------------------------
    # 链接导入
    # ------------------------------------------------------------------

    def import_links_file(self):
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "导入链接文件",
            "",
            "链接文件 (*.txt *.csv *.xlsx);;Excel文件 (*.xlsx);;文本文件 (*.txt);;CSV文件 (*.csv);;所有文件 (*.*)",
        )

        if not file_path:
            return

        try:
            path = Path(file_path)
            suffix = path.suffix.lower()

            if suffix == ".xlsx":
                imported_urls = self._read_excel_links(path)
            else:
                try:
                    text = path.read_text(encoding="utf-8")
                except UnicodeDecodeError:
                    text = path.read_text(encoding="gbk", errors="ignore")

                imported_urls = self._extract_and_simplify_urls_from_text(text)

            if not imported_urls:
                QMessageBox.warning(self, "导入失败", "文件中未识别到有效商品链接。")
                return

            current_urls = self._get_urls(update_input=False)

            merged = []
            seen = set()

            for url in current_urls + imported_urls:
                if url in seen:
                    continue

                seen.add(url)
                merged.append(url)

            added_count = max(len(merged) - len(current_urls), 0)
            duplicate_count = max(len(imported_urls) - added_count, 0)

            self.url_input.setPlainText("\n".join(merged))

            self.log(
                f"导入链接文件成功：{path.name}，"
                f"识别链接 {len(imported_urls)} 个，"
                f"新增 {added_count} 个，"
                f"重复 {duplicate_count} 个，"
                f"当前共 {len(merged)} 个。"
            )

        except ImportError:
            QMessageBox.critical(
                self,
                "导入失败",
                "导入 Excel 文件需要安装 openpyxl。\n\n请执行：\npip install openpyxl==3.1.5",
            )

        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"导入链接文件失败：{e}")

    def _read_excel_links(self, path: Path) -> list[str]:
        from openpyxl import load_workbook

        workbook = load_workbook(
            filename=path,
            read_only=True,
            data_only=True,
        )

        all_urls = []

        try:
            for sheet in workbook.worksheets:
                urls = self._read_links_from_worksheet(sheet)
                all_urls.extend(urls)

        finally:
            try:
                workbook.close()
            except Exception:
                pass

        return self._dedupe_url_list(all_urls)

    def _read_links_from_worksheet(self, sheet) -> list[str]:
        result = []

        url_column_index, header_row_index = self._find_excel_url_column(sheet)

        if url_column_index is not None:
            for row in sheet.iter_rows(
                min_row=header_row_index + 1,
                values_only=True,
            ):
                if not row:
                    continue

                if url_column_index >= len(row):
                    continue

                value = row[url_column_index]

                if value is None:
                    continue

                text = str(value).strip()

                if not text:
                    continue

                urls = self._extract_and_simplify_urls_from_text(text)
                result.extend(urls)

            return self._dedupe_url_list(result)

        max_scan_rows = 10000
        scanned_rows = 0

        for row in sheet.iter_rows(values_only=True):
            scanned_rows += 1

            if scanned_rows > max_scan_rows:
                break

            if not row:
                continue

            for value in row:
                if value is None:
                    continue

                text = str(value).strip()

                if not text:
                    continue

                urls = self._extract_and_simplify_urls_from_text(text)
                result.extend(urls)

        return self._dedupe_url_list(result)

    def _find_excel_url_column(self, sheet) -> tuple[int | None, int]:
        header_keywords = {
            "商品链接",
            "链接",
            "url",
            "商品url",
            "商品地址",
            "商品链接地址",
            "地址",
            "link",
            "链接地址",
            "商品页面",
            "商品网址",
        }

        max_header_scan_rows = 10

        for row_index, row in enumerate(
            sheet.iter_rows(
                min_row=1,
                max_row=max_header_scan_rows,
                values_only=True,
            ),
            start=1,
        ):
            if not row:
                continue

            for column_index, value in enumerate(row):
                if value is None:
                    continue

                header_text = self._normalize_excel_header(str(value))

                if not header_text:
                    continue

                if header_text in header_keywords:
                    return column_index, row_index

                if "url" in header_text:
                    return column_index, row_index

                if "链接" in header_text:
                    return column_index, row_index

                if "网址" in header_text:
                    return column_index, row_index

                if "地址" in header_text and "商品" in header_text:
                    return column_index, row_index

        return None, 1

    def _normalize_excel_header(self, text: str) -> str:
        if not text:
            return ""

        text = str(text).strip().lower()

        text = text.replace(" ", "")
        text = text.replace("\t", "")
        text = text.replace("\n", "")
        text = text.replace("\r", "")
        text = text.replace("_", "")
        text = text.replace("-", "")
        text = text.replace("：", "")
        text = text.replace(":", "")

        return text

    def _dedupe_url_list(self, urls: list[str]) -> list[str]:
        result = []
        seen = set()

        for url in urls:
            if not url:
                continue

            if url in seen:
                continue

            seen.add(url)
            result.append(url)

        return result

    # ------------------------------------------------------------------
    # 解析
    # ------------------------------------------------------------------

    def parse_product(self):
        urls = self._get_urls(update_input=True)

        if not urls:
            QMessageBox.warning(self, "提示", "请先输入商品链接。")
            return

        self.save_current_config()

        self.product = None
        self.products = []
        self.original_products = []
        self.failed_parse_items = []
        self.selection_file_path = None
        self.last_retry_items = []

        self.last_product_dir = None
        self.last_base_dir = None
        self.last_preview_path = None

        self.download_button.setEnabled(False)
        self.retry_failed_button.setEnabled(False)
        self.preview_button.setEnabled(False)
        self.import_selection_button.setEnabled(False)
        self.open_dir_button.setEnabled(False)
        self.progress_bar.setValue(0)
        self.clear_product_info()

        if len(urls) == 1:
            self.log("开始解析任务...")
        else:
            self.log(f"开始批量解析任务，共 {len(urls)} 个链接...")

        self.parse_button.setEnabled(False)
        self.download_button.setEnabled(False)
        self.retry_failed_button.setEnabled(False)
        self.preview_button.setEnabled(False)
        self.import_selection_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.import_links_button.setEnabled(False)

        self.parse_worker = BatchParseWorker(
            urls=urls,
            headless=self.headless_checkbox.isChecked(),
            login_wait_seconds=self.login_wait_spinbox.value(),
        )
        self.parse_worker.log_signal.connect(self.log)
        self.parse_worker.progress_signal.connect(self.progress_bar.setValue)
        self.parse_worker.success_signal.connect(self.on_parse_success)
        self.parse_worker.error_signal.connect(self.on_parse_error)
        self.parse_worker.stopped_signal.connect(self.on_parse_stopped)
        self.parse_worker.finished.connect(lambda: self.parse_button.setEnabled(True))
        self.parse_worker.start()

    def on_parse_success(self, payload):
        if isinstance(payload, dict):
            self.products = payload.get("products", []) or []
            self.failed_parse_items = payload.get("failed", []) or []
        else:
            self.products = payload or []
            self.failed_parse_items = []

        self.original_products = deepcopy(self.products)
        self.selection_file_path = None

        self.product = self.products[0] if self.products else None

        if not self.products:
            self.download_button.setEnabled(False)
            self.retry_failed_button.setEnabled(False)
            self.preview_button.setEnabled(False)
            self.import_selection_button.setEnabled(False)
            self.stop_button.setEnabled(False)
            self.import_links_button.setEnabled(True)
            return

        self._refresh_product_info_display()

        if len(self.products) == 1:
            if self.failed_parse_items:
                self.log(f"解析完成：成功 1 个，失败 {len(self.failed_parse_items)} 个。")
            else:
                self.log("商品解析完成，可以开始下载。")
        else:
            total_main = sum(len(p.main_images) for p in self.products)
            total_detail = sum(len(p.detail_images) for p in self.products)
            total_sku = sum(len(p.sku_images) for p in self.products)

            self.log(
                f"批量解析完成：成功 {len(self.products)} 个商品，"
                f"失败 {len(self.failed_parse_items)} 个链接，"
                f"主图 {total_main} 张，详情图 {total_detail} 张，SKU图 {total_sku} 张。"
            )

        self.download_button.setEnabled(True)
        self.retry_failed_button.setEnabled(bool(self.last_retry_items))
        self.preview_button.setEnabled(True)
        self.import_selection_button.setEnabled(True)
        self.stop_button.setEnabled(False)
        self.import_links_button.setEnabled(True)
        self.progress_bar.setValue(100)

    def on_parse_stopped(self, payload):
        if isinstance(payload, dict):
            self.products = payload.get("products", []) or []
            self.failed_parse_items = payload.get("failed", []) or []
        else:
            self.products = []
            self.failed_parse_items = []

        self.original_products = deepcopy(self.products)
        self.selection_file_path = None

        self.product = self.products[0] if self.products else None

        if self.products:
            self._refresh_product_info_display()

            self.download_button.setEnabled(True)
            self.retry_failed_button.setEnabled(bool(self.last_retry_items))
            self.preview_button.setEnabled(True)
            self.import_selection_button.setEnabled(True)

            self.log(f"解析任务已停止，已保留 {len(self.products)} 个已解析商品，可继续下载或预览。")
        else:
            self.download_button.setEnabled(False)
            self.retry_failed_button.setEnabled(False)
            self.preview_button.setEnabled(False)
            self.import_selection_button.setEnabled(False)
            self.log("解析任务已停止，未产生可下载商品。")

        self.stop_button.setEnabled(False)
        self.parse_button.setEnabled(True)
        self.import_links_button.setEnabled(True)

    def on_parse_error(self, message: str):
        self.log(message)

        self.stop_button.setEnabled(False)
        self.parse_button.setEnabled(True)
        self.import_links_button.setEnabled(True)
        self.preview_button.setEnabled(False)
        self.import_selection_button.setEnabled(False)
        self.retry_failed_button.setEnabled(bool(self.last_retry_items))

        QMessageBox.critical(self, "解析失败", message)

    # ------------------------------------------------------------------
    # 选择结果导入
    # ------------------------------------------------------------------

    def import_selection_file(self):
        if not self.original_products:
            QMessageBox.warning(self, "提示", "请先解析商品并生成预览。")
            return

        file_path, _ = QFileDialog.getOpenFileName(
            self,
            "导入选择结果",
            "",
            "选择结果文件 (*.json);;所有文件 (*.*)",
        )

        if not file_path:
            return

        try:
            path = Path(file_path)

            data = json.loads(path.read_text(encoding="utf-8"))

            if not isinstance(data, dict):
                QMessageBox.warning(self, "导入失败", "选择结果文件格式不正确。")
                return

            products_data = data.get("products", [])

            if not isinstance(products_data, list) or not products_data:
                QMessageBox.warning(self, "导入失败", "选择结果中没有商品数据。")
                return

            filtered_products, stats = self._filter_products_by_selection(products_data)

            if stats["matched_products"] == 0:
                QMessageBox.warning(
                    self,
                    "导入失败",
                    "选择结果与当前解析商品不匹配。\n\n"
                    "请确认 selected_images.json 是由当前商品预览页导出的。",
                )
                return

            if stats["after_total"] == 0:
                reply = QMessageBox.question(
                    self,
                    "确认导入",
                    "选择结果中没有任何已选图片。\n\n"
                    "导入后将没有可下载图片，是否继续？",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No,
                )

                if reply != QMessageBox.Yes:
                    return

            self.products = filtered_products
            self.product = self.products[0] if self.products else None
            self.selection_file_path = path

            self._refresh_product_info_display()

            self.download_button.setEnabled(stats["after_total"] > 0)
            self.preview_button.setEnabled(True if self.products else False)

            self.log(
                f"选择结果导入成功：{path.name}，"
                f"匹配商品 {stats['matched_products']} 个，"
                f"图片数量 {stats['before_total']} -> {stats['after_total']}。"
            )

            QMessageBox.information(
                self,
                "导入成功",
                f"选择结果导入成功！\n\n"
                f"匹配商品：{stats['matched_products']} 个\n"
                f"导入前图片：{stats['before_total']} 张\n"
                f"导入后图片：{stats['after_total']} 张\n\n"
                f"现在点击“开始下载”，将只下载选中的图片。",
            )

        except json.JSONDecodeError:
            QMessageBox.critical(self, "导入失败", "JSON 文件解析失败，请检查文件格式。")

        except Exception as e:
            QMessageBox.critical(self, "导入失败", f"导入选择结果失败：{e}")

    def _filter_products_by_selection(self, selected_products_data: list[dict]):
        source_products = deepcopy(self.original_products)

        stats = {
            "matched_products": 0,
            "before_total": sum(product.total_count() for product in source_products),
            "after_total": 0,
        }

        filtered_products = []

        for product_index, product in enumerate(source_products, start=1):
            selection_data = self._find_selection_for_product(
                product=product,
                product_index=product_index,
                selected_products_data=selected_products_data,
            )

            if not selection_data:
                filtered_products.append(product)
                stats["after_total"] += product.total_count()
                continue

            stats["matched_products"] += 1

            selected_sets = self._extract_selected_url_sets(selection_data)

            product.main_images = [
                img for img in product.main_images
                if img.url in selected_sets["main"]
            ]

            product.detail_images = [
                img for img in product.detail_images
                if img.url in selected_sets["detail"]
            ]

            product.sku_images = [
                img for img in product.sku_images
                if img.url in selected_sets["sku"]
            ]

            stats["after_total"] += product.total_count()
            filtered_products.append(product)

        return filtered_products, stats

    def _find_selection_for_product(
        self,
        product,
        product_index: int,
        selected_products_data: list[dict],
    ) -> dict | None:
        product_platform = (product.platform or "").strip()
        product_id = (product.product_id or "").strip()
        product_url = (product.url or "").strip()

        if product_platform and product_id:
            for item in selected_products_data:
                item_platform = str(item.get("platform", "")).strip()
                item_product_id = str(item.get("product_id", "")).strip()

                if item_platform == product_platform and item_product_id == product_id:
                    return item

        if product_url:
            for item in selected_products_data:
                item_url = str(item.get("url", "")).strip()

                if item_url == product_url:
                    return item

        for item in selected_products_data:
            try:
                item_index = int(item.get("index", 0))
            except Exception:
                item_index = 0

            if item_index == product_index:
                return item

        return None

    def _extract_selected_url_sets(self, selection_data: dict) -> dict[str, set[str]]:
        selected = selection_data.get("selected", {}) or {}

        return {
            "main": self._extract_url_set_from_selected_items(selected.get("main", [])),
            "detail": self._extract_url_set_from_selected_items(selected.get("detail", [])),
            "sku": self._extract_url_set_from_selected_items(selected.get("sku", [])),
        }

    def _extract_url_set_from_selected_items(self, items) -> set[str]:
        result = set()

        if not isinstance(items, list):
            return result

        for item in items:
            if isinstance(item, dict):
                url = str(item.get("url", "")).strip()
            else:
                url = str(item).strip()

            if url:
                result.add(url)

        return result

    # ------------------------------------------------------------------
    # 图片预览
    # ------------------------------------------------------------------

    def preview_images(self):
        if not self.products:
            QMessageBox.warning(self, "提示", "请先解析商品。")
            return

        base_dir = self.path_input.text().strip()

        if not base_dir:
            QMessageBox.warning(self, "提示", "请选择保存路径。")
            return

        self.save_current_config()

        try:
            Path(base_dir).mkdir(parents=True, exist_ok=True)

            preview_path = PreviewGenerator.save_preview(
                base_dir=base_dir,
                products=self.products,
            )

            self.last_preview_path = preview_path

            self.log(f"图片预览已生成：{preview_path}")

            os.startfile(str(preview_path))

        except Exception as e:
            QMessageBox.critical(self, "预览失败", f"生成图片预览失败：{e}")

    # ------------------------------------------------------------------
    # 下载
    # ------------------------------------------------------------------

    def download_images(self):
        if not self.products:
            QMessageBox.warning(self, "提示", "请先解析商品。")
            return

        base_dir = self.path_input.text().strip()
        if not base_dir:
            QMessageBox.warning(self, "提示", "请选择保存路径。")
            return

        self.save_current_config()

        selected_types = {
            "main": self.main_checkbox.isChecked(),
            "detail": self.detail_checkbox.isChecked(),
            "sku": self.sku_checkbox.isChecked(),
        }

        high_quality = self.high_quality_checkbox.isChecked()

        if not any(selected_types.values()):
            QMessageBox.warning(self, "提示", "请至少选择一种需要下载的图片类型。")
            return

        total_selected_images = 0
        for product in self.products:
            if selected_types["main"]:
                total_selected_images += len(product.main_images)
            if selected_types["detail"]:
                total_selected_images += len(product.detail_images)
            if selected_types["sku"]:
                total_selected_images += len(product.sku_images)

        if total_selected_images == 0:
            QMessageBox.warning(
                self,
                "提示",
                "当前选择条件下没有可下载图片。\n\n"
                "请检查：\n"
                "1. 是否导入了空的选择结果；\n"
                "2. 是否取消了对应图片类型；\n"
                "3. 是否需要重新解析商品。",
            )
            return

        Path(base_dir).mkdir(parents=True, exist_ok=True)

        self.progress_bar.setValue(0)
        self.download_button.setEnabled(False)
        self.retry_failed_button.setEnabled(False)
        self.parse_button.setEnabled(False)
        self.preview_button.setEnabled(False)
        self.import_selection_button.setEnabled(False)
        self.stop_button.setEnabled(True)
        self.import_links_button.setEnabled(False)

        if len(self.products) == 1:
            self.log("开始下载任务...")
        else:
            self.log(f"开始批量下载任务，共 {len(self.products)} 个商品...")

        if self.selection_file_path:
            self.log(f"已应用选择结果：{self.selection_file_path}")

        self.download_worker = BatchDownloadWorker(
            products=self.products,
            base_dir=base_dir,
            selected_types=selected_types,
            failed_parse_items=self.failed_parse_items,
            high_quality=high_quality,
            download_timeout=self.timeout_spinbox.value(),
            download_retries=self.retries_spinbox.value(),
            organize_by_date=self.organize_by_date_checkbox.isChecked(),
            organize_by_platform=self.organize_by_platform_checkbox.isChecked(),
            dedupe_images=self.dedupe_images_checkbox.isChecked(),
            image_output_format=self._get_image_format_combo_value(),
            filter_small_images_enabled=self.filter_small_images_checkbox.isChecked(),
            min_image_width=self.min_width_spinbox.value(),
            min_image_height=self.min_height_spinbox.value(),
        )

        self.download_worker.log_signal.connect(self.log)
        self.download_worker.progress_signal.connect(self.progress_bar.setValue)
        self.download_worker.success_signal.connect(self.on_download_success)
        self.download_worker.error_signal.connect(self.on_download_error)
        self.download_worker.stopped_signal.connect(self.on_download_stopped)
        self.download_worker.finished.connect(self.on_download_finished)
        self.download_worker.start()

    def on_download_success(self, base_dir, last_product_dir, payload):
        self.last_base_dir = base_dir
        self.last_product_dir = last_product_dir
        self.open_dir_button.setEnabled(True)

        if isinstance(payload, dict):
            result = payload.get("result")
            self.last_retry_items = payload.get("retry_items", []) or []
        else:
            result = payload
            self.last_retry_items = []

        self.retry_failed_button.setEnabled(bool(self.last_retry_items))

        retry_text = ""
        if self.last_retry_items:
            retry_text = f"\n失败图片可重试：{len(self.last_retry_items)} 张"

        QMessageBox.information(
            self,
            "下载完成",
            f"下载完成！\n"
            f"计划下载：{result.total} 张\n"
            f"成功：{result.success} 张\n"
            f"失败：{result.failed} 张\n"
            f"成功率：{result.success_rate}%\n"
            f"MD5去重处理：{result.duplicate_removed} 张\n"
            f"格式转换成功：{result.converted_count} 张\n"
            f"格式转换失败：{result.convert_failed} 张\n"
            f"小图过滤：{result.small_filtered_count} 张"
            f"{retry_text}",
        )

    def on_download_stopped(self, payload):
        if isinstance(payload, dict):
            self.last_base_dir = payload.get("base_dir")
            self.last_product_dir = payload.get("last_product_dir")
            result = payload.get("result")
            self.last_retry_items = payload.get("retry_items", []) or []
        else:
            result = None
            self.last_retry_items = []

        self.open_dir_button.setEnabled(True)
        self.preview_button.setEnabled(True if self.products else False)
        self.import_selection_button.setEnabled(True if self.original_products else False)
        self.stop_button.setEnabled(False)
        self.parse_button.setEnabled(True)
        self.download_button.setEnabled(True if self.products else False)
        self.retry_failed_button.setEnabled(bool(self.last_retry_items))
        self.import_links_button.setEnabled(True)

        if result:
            self.log(
                f"下载任务已停止：计划 {result.total} 张，"
                f"已成功 {result.success} 张，失败 {result.failed} 张。"
            )
            QMessageBox.information(
                self,
                "任务已停止",
                f"下载任务已停止。\n"
                f"计划下载：{result.total} 张\n"
                f"已成功：{result.success} 张\n"
                f"失败：{result.failed} 张\n"
                f"已生成下载报告和失败清单。",
            )
        else:
            self.log("下载任务已停止。")

    def on_download_error(self, message: str):
        self.log(message)

        self.stop_button.setEnabled(False)
        self.parse_button.setEnabled(True)
        self.download_button.setEnabled(True if self.products else False)
        self.retry_failed_button.setEnabled(bool(self.last_retry_items))
        self.preview_button.setEnabled(True if self.products else False)
        self.import_selection_button.setEnabled(True if self.original_products else False)
        self.import_links_button.setEnabled(True)

        QMessageBox.critical(self, "下载失败", message)

    def on_download_finished(self):
        self.parse_button.setEnabled(True)
        self.download_button.setEnabled(True if self.products else False)
        self.retry_failed_button.setEnabled(bool(self.last_retry_items))
        self.preview_button.setEnabled(True if self.products else False)
        self.import_selection_button.setEnabled(True if self.original_products else False)
        self.stop_button.setEnabled(False)
        self.import_links_button.setEnabled(True)

    # ------------------------------------------------------------------
    # 失败重试
    # ------------------------------------------------------------------

    def retry_failed_images(self):
        if not self.last_retry_items:
            QMessageBox.information(self, "提示", "当前没有可重试的失败图片。")
            return

        if self._is_task_running():
            QMessageBox.warning(
                self,
                "提示",
                "当前有解析或下载任务正在运行，请等待任务结束后再重试失败图片。",
            )
            return

        reply = QMessageBox.question(
            self,
            "确认重试",
            f"当前有 {len(self.last_retry_items)} 张失败图片可重试。\n\n"
            f"是否开始重试？",
            QMessageBox.Yes | QMessageBox.No,
            QMessageBox.Yes,
        )

        if reply != QMessageBox.Yes:
            return

        self.progress_bar.setValue(0)

        self.parse_button.setEnabled(False)
        self.download_button.setEnabled(False)
        self.retry_failed_button.setEnabled(False)
        self.preview_button.setEnabled(False)
        self.import_selection_button.setEnabled(False)
        self.import_links_button.setEnabled(False)
        self.stop_button.setEnabled(True)

        self.log(f"开始重试失败图片，共 {len(self.last_retry_items)} 张...")

        self.retry_worker = RetryFailedDownloadWorker(
            retry_items=self.last_retry_items,
            download_timeout=self.timeout_spinbox.value(),
            download_retries=self.retries_spinbox.value(),
            high_quality=self.high_quality_checkbox.isChecked(),
        )

        self.retry_worker.log_signal.connect(self.log)
        self.retry_worker.progress_signal.connect(self.progress_bar.setValue)
        self.retry_worker.success_signal.connect(self.on_retry_failed_success)
        self.retry_worker.error_signal.connect(self.on_retry_failed_error)
        self.retry_worker.stopped_signal.connect(self.on_retry_failed_stopped)
        self.retry_worker.finished.connect(self.on_retry_failed_finished)
        self.retry_worker.start()

    def on_retry_failed_success(self, payload):
        if isinstance(payload, dict):
            result = payload.get("result")
            self.last_retry_items = payload.get("retry_items", []) or []
        else:
            result = None
            self.last_retry_items = []

        self.retry_failed_button.setEnabled(bool(self.last_retry_items))
        self.open_dir_button.setEnabled(True)

        if result:
            QMessageBox.information(
                self,
                "重试完成",
                f"失败图片重试完成！\n\n"
                f"重试总数：{result.total} 张\n"
                f"成功：{result.success} 张\n"
                f"失败：{result.failed} 张\n"
                f"成功率：{result.success_rate}%\n\n"
                f"剩余可重试：{len(self.last_retry_items)} 张",
            )

    def on_retry_failed_stopped(self, payload):
        if isinstance(payload, dict):
            result = payload.get("result")
            self.last_retry_items = payload.get("retry_items", []) or []
        else:
            result = None
            self.last_retry_items = []

        self.retry_failed_button.setEnabled(bool(self.last_retry_items))

        if result:
            self.log(
                f"失败图片重试已停止：成功 {result.success} 张，失败 {result.failed} 张。"
            )

    def on_retry_failed_error(self, message: str):
        self.log(message)
        self.retry_failed_button.setEnabled(bool(self.last_retry_items))
        QMessageBox.critical(self, "重试失败", message)

    def on_retry_failed_finished(self):
        self.parse_button.setEnabled(True)
        self.download_button.setEnabled(True if self.products else False)
        self.retry_failed_button.setEnabled(bool(self.last_retry_items))
        self.preview_button.setEnabled(True if self.products else False)
        self.import_selection_button.setEnabled(True if self.original_products else False)
        self.import_links_button.setEnabled(True)
        self.stop_button.setEnabled(False)

    # ------------------------------------------------------------------
    # 目录
    # ------------------------------------------------------------------

    def open_last_dir(self):
        if self.last_product_dir and Path(self.last_product_dir).exists():
            os.startfile(str(self.last_product_dir))
            return

        if self.last_base_dir and Path(self.last_base_dir).exists():
            os.startfile(str(self.last_base_dir))
            return

    # ------------------------------------------------------------------
    # 停止任务
    # ------------------------------------------------------------------

    def stop_current_task(self):
        stopped = False

        if self.parse_worker and self.parse_worker.isRunning():
            if hasattr(self.parse_worker, "stop"):
                self.parse_worker.stop()
                stopped = True

        if self.download_worker and self.download_worker.isRunning():
            if hasattr(self.download_worker, "stop"):
                self.download_worker.stop()
                stopped = True

        if self.retry_worker and self.retry_worker.isRunning():
            if hasattr(self.retry_worker, "stop"):
                self.retry_worker.stop()
                stopped = True

        if stopped:
            self.log("正在停止任务，请稍候...")
            self.stop_button.setEnabled(False)
        else:
            self.log("当前没有正在运行的任务。")

    # ------------------------------------------------------------------
    # 日志
    # ------------------------------------------------------------------

    def clear_log(self):
        self.log_text.clear()

    def log(self, message: str):
        now = datetime.now().strftime("%H:%M:%S")
        self.log_text.append(f"[{now}] {message}")

    # ------------------------------------------------------------------
    # 工具方法
    # ------------------------------------------------------------------

    def clear_product_info(self):
        self.platform_label.setText("平台：-")
        self.product_id_label.setText("商品ID：-")
        self.title_label.setText("商品标题：-")
        self.main_count_label.setText("主图：- 张")
        self.detail_count_label.setText("详情图：- 张")
        self.sku_count_label.setText("SKU图：- 张")

    def _refresh_product_info_display(self):
        if not self.products:
            self.clear_product_info()
            return

        if len(self.products) == 1:
            product = self.products[0]

            self.platform_label.setText(f"平台：{product.platform}")
            self.product_id_label.setText(f"商品ID：{product.product_id}")
            self.title_label.setText(f"商品标题：{product.title}")

            self.main_count_label.setText(f"主图：{len(product.main_images)} 张")
            self.detail_count_label.setText(f"详情图：{len(product.detail_images)} 张")
            self.sku_count_label.setText(f"SKU图：{len(product.sku_images)} 张")

            return

        total_main = sum(len(p.main_images) for p in self.products)
        total_detail = sum(len(p.detail_images) for p in self.products)
        total_sku = sum(len(p.sku_images) for p in self.products)

        platforms = sorted(set(p.platform for p in self.products))

        self.platform_label.setText(f"平台：批量 / {', '.join(platforms)}")
        self.product_id_label.setText(f"商品ID：共 {len(self.products)} 个商品")
        self.title_label.setText("商品标题：批量任务")

        self.main_count_label.setText(f"主图：{total_main} 张")
        self.detail_count_label.setText(f"详情图：{total_detail} 张")
        self.sku_count_label.setText(f"SKU图：{total_sku} 张")

    def _get_image_format_combo_value(self) -> str:
        if not hasattr(self, "image_format_combo"):
            return "original"

        value = self.image_format_combo.currentData()

        if value in ["original", "jpg", "png", "webp"]:
            return value

        return "original"

    def _set_image_format_combo_value(self, value: str) -> None:
        if not hasattr(self, "image_format_combo"):
            return

        value = value or "original"

        for index in range(self.image_format_combo.count()):
            if self.image_format_combo.itemData(index) == value:
                self.image_format_combo.setCurrentIndex(index)
                return

        self.image_format_combo.setCurrentIndex(0)

    def _get_urls(self, update_input: bool = True) -> list[str]:
        text = self.url_input.toPlainText().strip()

        if not text:
            return []

        result = self._extract_and_simplify_urls_from_text(text)

        if update_input and result:
            new_text = "\n".join(result)

            if new_text != text:
                self.url_input.setPlainText(new_text)

        return result

    def _extract_and_simplify_urls_from_text(self, text: str) -> list[str]:
        if not text:
            return []

        text = text.replace("\r\n", "\n").replace("\r", "\n")

        raw_urls = re.findall(
            r"https?://[^\s<>'\"，,；;。]+",
            text,
            flags=re.I,
        )

        if not raw_urls:
            raw_urls = [line.strip() for line in text.splitlines() if line.strip()]

        result = []
        seen = set()

        for raw_url in raw_urls:
            cleaned = self._clean_input_url(raw_url)
            simplified = self._simplify_product_url(cleaned)

            if not simplified:
                continue

            if simplified in seen:
                continue

            seen.add(simplified)
            result.append(simplified)

        return result

    def _clean_input_url(self, url: str) -> str:
        if not url:
            return ""

        url = str(url).strip()

        url = url.strip(" \t\r\n'\"<>")
        url = url.rstrip("，,；;。.)）]】")

        if not re.match(r"^https?://", url, flags=re.I):
            return ""

        return url

    def _simplify_product_url(self, url: str) -> str:
        if not url:
            return ""

        try:
            parsed = urlparse(url)
            host = parsed.netloc.lower()
            query = parse_qs(parsed.query)

            if "jd.com" in host:
                match = re.search(r"/(\d+)\.html", url)

                if match:
                    sku_id = match.group(1)
                    return f"https://item.jd.com/{sku_id}.html"

                return url

            if "taobao.com" in host:
                item_id = query.get("id", [""])[0]

                if item_id:
                    return f"https://item.taobao.com/item.htm?id={item_id}"

                return url

            if "tmall.com" in host:
                item_id = query.get("id", [""])[0]

                if item_id:
                    return f"https://detail.tmall.com/item.htm?id={item_id}"

                return url
                
            if "1688.com" in host:
                match = re.search(r"/offer/(\d+)\.html", url)
                if match:
                    offer_id = match.group(1)
                    return f"https://detail.1688.com/offer/{offer_id}.html"

                offer_id = query.get("offerId", [""])[0] or query.get("offerid", [""])[0]
                if offer_id:
                    return f"https://detail.1688.com/offer/{offer_id}.html"

                return url

            if "pinduoduo.com" in host or "yangkeduo.com" in host:
                goods_id = query.get("goods_id", [""])[0]

                if goods_id:
                    return f"https://mobile.yangkeduo.com/goods.html?goods_id={goods_id}"

                return url

            return url

        except Exception:
            return url
    
    def append_log(self, message: str):
        """
        追加日志到实时日志框。

        兼容当前 main_window.py 中不同日志控件命名。
        如果找不到明确日志控件，则自动寻找除 url_input 外的 QTextEdit。
        """
        try:
            text = f"[{datetime.now().strftime('%H:%M:%S')}] {message}"

            # 常见日志控件命名兼容
            candidate_names = [
                "log_output",
                "log_text",
                "log_edit",
                "log_browser",
                "log_view",
                "log_area",
                "log_box",
            ]

            for name in candidate_names:
                widget = getattr(self, name, None)
                if widget is not None and hasattr(widget, "append"):
                    widget.append(text)
                    return

            # 自动查找 QTextEdit，排除商品链接输入框
            for value in self.__dict__.values():
                if isinstance(value, QTextEdit) and value is not getattr(self, "url_input", None):
                    value.append(text)
                    return

            # 实在找不到日志框时，至少打印到控制台
            print(text)

        except Exception:
            try:
                print(message)
            except Exception:
                pass


    def refresh_resume_task_button(self):
        """
        刷新“继续上次任务”按钮状态。

        有未完成任务：
            按钮可点击并高亮。

        没有未完成任务：
            按钮禁用并灰色。
        """
        try:
            button = getattr(self, "resume_task_button", None)
            if button is None:
                return

            base_dir = self._get_current_save_dir_for_task_state()
            state_path = TaskStateManager.find_latest_unfinished(base_dir)

            if state_path:
                button.setEnabled(True)
                button.setToolTip(f"检测到未完成任务：{state_path}")
                button.setStyleSheet(
                    """
                    QPushButton {
                        background-color: #ff9800;
                        color: white;
                        font-weight: bold;
                    }
                    QPushButton:hover {
                        background-color: #ffa726;
                    }
                    """
                )
            else:
                button.setEnabled(False)
                button.setToolTip("没有未完成任务")
                button.setStyleSheet(
                    """
                    QPushButton {
                        background-color: #666666;
                        color: #cccccc;
                    }
                    """
                )

        except Exception:
            try:
                button = getattr(self, "resume_task_button", None)
                if button:
                    button.setEnabled(False)
            except Exception:
                pass

    
    def resume_last_task(self):
        """
        继续上次任务 - 2A 低风险版。

        功能：
            1. 查找 output/任务状态 下最近一个未完成任务；
            2. 提取 pending / failed / running 商品链接；
            3. 恢复到商品链接输入框；
            4. 用户手动点击“解析商品”继续。
        """
        try:
            if hasattr(self, "_is_task_running") and self._is_task_running():
                QMessageBox.warning(self, "任务运行中", "当前有任务正在运行，请先停止或等待完成。")
                return

            base_dir = self._get_current_save_dir_for_task_state()

            state_path = TaskStateManager.find_latest_unfinished(base_dir)

            if not state_path:
                QMessageBox.information(self, "继续上次任务", "没有找到未完成任务。")
                self.append_log("没有找到未完成任务。")
                return

            manager = TaskStateManager.load(state_path)
            urls = manager.get_unfinished_urls()

            if not urls:
                QMessageBox.information(self, "继续上次任务", "未完成任务中没有待恢复的商品链接。")
                self.append_log(f"未完成任务没有待恢复链接：{state_path}")
                return

            self.url_input.setPlainText("\n".join(urls))
            
            try:
                manager.mark_resumed()
                self.append_log(f"原任务状态已标记为已恢复：{state_path}")
            except Exception as e:
                self.append_log(f"原任务状态标记为已恢复失败：{e}")

            self.refresh_resume_task_button()


            # 清理当前解析结果，避免用户误以为还是旧商品
            self.product = None
            self.products = []
            self.original_products = []
            self.failed_parse_items = []
            self.selection_file_path = None

            self.platform_label.setText("平台：-")
            self.product_id_label.setText("商品ID：-")
            self.title_label.setText("商品标题：-")

            self.append_log(f"已恢复上次未完成任务：{state_path}")
            self.append_log(f"已恢复待处理商品链接：{len(urls)} 个")
            self.append_log("请点击“解析商品”，解析完成后再点击“开始下载”继续任务。")

            QMessageBox.information(
                self,
                "继续上次任务",
                f"已恢复 {len(urls)} 个未完成商品链接。\n\n"
                "请点击“解析商品”，解析完成后再点击“开始下载”。"
            )

        except Exception as e:
            QMessageBox.critical(self, "继续上次任务失败", str(e))
            self.append_log(f"继续上次任务失败：{e}")

    def _get_current_save_dir_for_task_state(self) -> str:
        """
        获取当前保存目录，用于查找 output/任务状态。

        兼容不同版本 main_window.py 里的保存路径控件命名。
        """
        candidates = [
            "save_dir_input",
            "save_path_input",
            "output_dir_input",
            "base_dir_input",
        ]

        for name in candidates:
            widget = getattr(self, name, None)
            if widget:
                try:
                    value = widget.text().strip()
                    if value:
                        return value
                except Exception:
                    pass

        try:
            value = getattr(self.config, "save_dir", "")
            if value:
                return value
        except Exception:
            pass

        return "output"

    def closeEvent(self, event):
        self.save_current_config()

        running_tasks = []

        if self.parse_worker and self.parse_worker.isRunning():
            running_tasks.append("解析任务")

        if self.download_worker and self.download_worker.isRunning():
            running_tasks.append("下载任务")

        if self.retry_worker and self.retry_worker.isRunning():
            running_tasks.append("失败图片重试")

        if self.login_browser_worker and self.login_browser_worker.isRunning():
            running_tasks.append("登录浏览器")

        if running_tasks:
            task_text = "、".join(running_tasks)

            reply = QMessageBox.question(
                self,
                "仍有任务运行中",
                f"当前仍有任务正在运行：{task_text}\n\n"
                f"如果正在使用登录浏览器，请先关闭浏览器窗口，"
                f"或点击“结束登录浏览器”。\n\n"
                f"确定要尝试停止任务并退出吗？",
                QMessageBox.Yes | QMessageBox.No,
                QMessageBox.No,
            )

            if reply != QMessageBox.Yes:
                event.ignore()
                return

            if self.parse_worker and self.parse_worker.isRunning():
                try:
                    self.parse_worker.stop()
                except Exception:
                    pass

            if self.download_worker and self.download_worker.isRunning():
                try:
                    self.download_worker.stop()
                except Exception:
                    pass

            if self.retry_worker and self.retry_worker.isRunning():
                try:
                    self.retry_worker.stop()
                except Exception:
                    pass

            if self.login_browser_worker and self.login_browser_worker.isRunning():
                try:
                    self.login_browser_worker.stop()
                except Exception:
                    pass

            still_running = []

            if self.parse_worker and self.parse_worker.isRunning():
                if not self.parse_worker.wait(3000):
                    still_running.append("解析任务")

            if self.download_worker and self.download_worker.isRunning():
                if not self.download_worker.wait(3000):
                    still_running.append("下载任务")

            if self.retry_worker and self.retry_worker.isRunning():
                if not self.retry_worker.wait(3000):
                    still_running.append("失败图片重试")

            if self.login_browser_worker and self.login_browser_worker.isRunning():
                if not self.login_browser_worker.wait(3000):
                    still_running.append("登录浏览器")

            if still_running:
                QMessageBox.warning(
                    self,
                    "无法立即退出",
                    "以下任务仍未结束：\n\n"
                    + "\n".join(still_running)
                    + "\n\n请先停止任务或关闭登录浏览器窗口后再退出。",
                )
                event.ignore()
                return

        QMainWindow.closeEvent(self, event)
